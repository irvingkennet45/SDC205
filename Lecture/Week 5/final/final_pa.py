# Kenneth Irving (kenirv5642)
# Date: 08-30-2026
# Description: Program containing functions to calculate numbers sum, append income data to CSV, generate an Excel pie chart with openpyxl, and create a vertical bar chart using matplotlib.

from datetime import datetime
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
import matplotlib.pyplot as plt

# Student ID constant; used in chart titles
STUDENT_ID = "kenirv5642"

# Asks the user for five numbers using a loop, adds up the total, and displays it.
def askUser():
    total = 0
    
    # Loop 5 times to prompt the user for a number, add each number to the running total, and display the accumulated sum.
    for i in range(1, 6):
        user_num = float(input(f"Enter number {i} of 5: "))
        total += user_num
        
    print(f"\nThe total of the five numbers is: {total}\n")

# Asks the user for names and annual incomes of five people and appends them to final.csv.
def askIncome():
    csv_file_path = Path(__file__).parent / "final.csv"
    
    # Open final.csv in append mode ('a') so existing entries are preserved and new rows are appended
    with open(csv_file_path, "a", newline="") as file:
        # Loop 5 times to prompt the user for each person's name and annual income, then write each entry as a new line in final.csv.
        for i in range(1, 6):
            name = input(f"Enter person {i}'s name: ").strip()
            income = input(f"Enter annual income for {name}: ").strip()
            # Append the entry followed by a newline character
            file.write(f"{name},{income}\n")
            
    print("\nSuccessfully appended 5 income records to final.csv.\n")


# Reads income data from final.csv, writes to final.xlsx, and generates a pie chart.
def excelPie():
    csv_file_path = Path(__file__).parent / "final.csv"
    excel_file_path = Path(__file__).parent / "final.xlsx"
    
    # Init Excel workbook
    wb = Workbook()
    
    # Get the currently active worksheet
    ws = wb.active
    
    # Set the worksheet title
    ws.title = "Income Data"
    
    # Add column headers to the worksheet
    ws.append(["Name", "Annual Income"])
    
    # Read data from final.csv and append to worksheet with income cast to int
    with open(csv_file_path, "r") as file:
        reader = csv.reader(file)
        for row in reader:
            if row and len(row) >= 2:
                name = row[0].strip()
                income = int(float(row[1].strip()))  # Cast numerical value to int so pie chart displays correctly
                ws.append([name, income])
                
    # Instantiate a new openpyxl PieChart object
    chart = PieChart()
    
    # Format today's date (e.g., November 4, 2022)
    today_date = datetime.now().strftime("%B %d, %Y")
    
    # Set the title of the pie chart to include StudentID and today's date
    chart.title = f"{STUDENT_ID} {today_date}"
    
    # Define reference range for numerical data (incomes in column 2, rows 1 to max_row)
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    
    # Define reference range for categories/labels (names in column 1, rows 2 to max_row)
    labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    
    # Add data to the chart specifying that column headers provide the data title
    chart.add_data(data, titles_from_data=True)
    
    # Set categories (names) for the slices of the pie chart
    chart.set_categories(labels)
    
    # Add the pie chart to worksheet cell D2
    ws.add_chart(chart, "D2")
    
    # Save workbook to the destination Excel file final.xlsx
    wb.save(excel_file_path)
    
    print(f"Excel pie chart successfully created and saved to '{excel_file_path}'.\n")

# Reads income data from final.csv and generates a vertical bar graph using matplotlib.
def verticalBar():
    csv_file_path = Path(__file__).parent / "final.csv"
    
    names = []
    incomes = []
    
    # Read the data from final.csv
    with open(csv_file_path, "r") as file:
        reader = csv.reader(file)
        for row in reader:
            if row and len(row) >= 2:
                names.append(row[0].strip())
                incomes.append(int(float(row[1].strip())))
                
    # Format today's date (e.g., November 4, 2022)
    today_date = datetime.now().strftime("%B %d, %Y")
    
    # Create matplotlib vertical bar chart
    plt.figure(figsize=(10, 6))
    plt.bar(names, incomes, color="skyblue", edgecolor="navy")
    plt.title(f"{STUDENT_ID} {today_date}", fontsize=14, fontweight="bold")
    plt.xlabel("Name", fontsize=12)
    plt.ylabel("Annual Income ($)", fontsize=12)
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    plt.show()

# Define the main function to call the other functions in sequence
if __name__ == "__main__":
    print("=" * 60)
    print("1. Testing askUser()...")
    print("=" * 60)
    askUser()
    
    print("=" * 60)
    print("2. Testing askIncome()...")
    print("=" * 60)
    askIncome()
    
    print("=" * 60)
    print("3. Testing excelPie()...")
    print("=" * 60)
    excelPie()
    
    print("=" * 60)
    print("4. Testing verticalBar()...")
    print("=" * 60)
    verticalBar()

